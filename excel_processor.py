# excel_processor.py

import pandas as pd
import openpyxl
import win32com.client as win32


class DataProcessor:
    """Base class for processing data files and writing them to Excel sheets."""

    def __init__(self, file_path):
        self.file_path = file_path

    def process(self, sheet):
        raise NotImplementedError("Subclasses should implement this method")


class Data1Processor(DataProcessor):
    """Processes data from a text file and writes it to the Excel sheet."""

    def process(self, sheet):
        with open(self.file_path, "r", encoding="utf-8") as file:
            data = file.readlines()
        for row_idx, line in enumerate(data, start=3):
            values = line.strip().split("\t")
            for col_idx, value in enumerate(values, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)


class Data2Processor(DataProcessor):
    """Processes modified data from a text file and writes it to the Excel sheet."""

    def process(self, sheet):
        modified_data = self.modify_data2()
        last_row = self.get_last_row(sheet) + 1
        for row_idx, line in enumerate(modified_data, start=last_row):
            values = line.strip().split("\t")
            for col_idx, value in enumerate(values, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)


def modify_data2(self):
    """
    Modifies data by inserting an extra column in the data.
    Handles IndexError when values list does not have the expected length.
    """
    modified_data = []
    try:
        with open(self.file_path, "r", encoding="utf-8") as file:
            lines = file.readlines()

        if not lines:
            raise ValueError("Data file is empty.")

        header = lines[0].strip().split("\t")

        if "Pu" not in header:
            raise ValueError("Column 'Pu' not found in header.")

        pu_index = header.index("Pu")
        header.insert(pu_index + 1, header[pu_index])  # Duplicate 'Pu' column
        modified_data.append("\t".join(header))

        for line in lines[1:]:
            values = line.strip().split("\t")

            # Check if values list is long enough to access pu_index
            if len(values) <= pu_index:
                print(f"Warning: Skipping line due to insufficient columns: {line}")
                continue  # Skip this line

            values.insert(pu_index + 1, values[pu_index])
            modified_data.append("\t".join(values))

    except FileNotFoundError:
        print(f"Error: File not found at {self.file_path}")
    except ValueError as e:
        print(f"Error: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

    return modified_data

    def get_last_row(self, sheet):
        """Finds the last row with data in the Excel sheet."""
        last_row = sheet.max_row
        while last_row > 0 and all(
            sheet.cell(row=last_row, column=col).value is None
            for col in range(1, sheet.max_column + 1)
        ):
            last_row -= 1
        return last_row


class Data3Processor(DataProcessor):
    """Processes data from the third data file and writes it to the Excel sheet."""

    def process(self, sheet):
        with open(self.file_path, "r", encoding="utf-8") as file:
            data = file.readlines()
        for row_idx, line in enumerate(data, start=1):
            values = line.strip().split("\t")
            cleaned_values = [self.clean_decimal(value) for value in values]
            for col_idx, value in enumerate(cleaned_values, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

    @staticmethod
    def clean_decimal(value):
        """Cleans decimal values to remove unnecessary zeros."""
        try:
            value = str(value)
            float_value = float(value)
            if float_value.is_integer():
                return int(float_value)
            return float(f"{float_value:.4f}".rstrip("0").rstrip("."))
        except ValueError:
            return value


def create_temp_excel(data1_path, data2_path, data3_path, excel_path, temp_excel_path):
    """
    Creates a temporary Excel file by processing data from text files and
    writing the results into specified sheets.

    :param data1_path: Path to the first data file (text file)
    :param data2_path: Path to the second data file (text file)
    :param data3_path: Path to the third data file (text file)
    :param excel_path: Path to the original Excel file
    :param temp_excel_path: Path to save the temporary Excel file
    """
    wb = openpyxl.load_workbook(excel_path, keep_vba=True)
    code_sheet = wb["code"]
    Data1Processor(data1_path).process(code_sheet)
    Data2Processor(data2_path).process(code_sheet)

    section_sheet = wb["section"]
    Data3Processor(data3_path).process(section_sheet)

    wb.save(temp_excel_path)
    wb.close()

    # Refresh Excel calculations and save
    excel = win32.Dispatch("Excel.Application")
    wb = excel.Workbooks.Open(temp_excel_path)
    excel.Calculate()
    wb.Save()
    wb.Close()
    excel.Quit()


def set_clipboard_from_excel(file_path):
    """
    Copies data from the first sheet of the specified Excel file to the clipboard.

    :param file_path: Path to the Excel file
    """
    excel = win32.Dispatch("Excel.Application")
    workbook = excel.Workbooks.Open(file_path)
    worksheet = workbook.Worksheets(1)
    worksheet.UsedRange.Copy()
    workbook.Close(False)
    excel.Quit()
    print("Data copied to clipboard from Excel file.")
